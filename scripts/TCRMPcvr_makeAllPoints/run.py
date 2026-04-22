#!/usr/bin/env python3
"""
TCRMPcvr_makeAllPoints/run.py

Parse all TCRMP CVR spreadsheets (2001-2025) into a unified CSV.
Handles three format generations: legacy (.xls), intermediate, and modern (.xlsx).

Usage:
    python run.py                              # uses default paths
    python run.py /path/to/TCRMP_CVR           # custom input dir
    python run.py /path/to/TCRMP_CVR ./out     # custom input + output dirs

Outputs:
    output/all_points_YYYYMMDD.csv   - Every benthic point observation
    output/master_codes.csv          - Species code lookup table
    output/parse_log.txt             - Parse log with errors and stats
"""
import os
import re
import sys
import warnings
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl
import xlrd
import pandas as pd

warnings.filterwarnings('ignore')

# Excel error values to filter (cached formula errors read by openpyxl)
EXCEL_ERRORS = {'#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NUM!', '#NAME?', '#NULL!'}

def _clean_species(val):
    """Return cleaned species code string, or None if invalid."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if not s or s in EXCEL_ERRORS or s == '0':
        return None
    return s

# ---------------------------------------------------------------------------
# Default paths. In the seg_AI_img_full_april2026 layout, the 839 TCRMP_CVR
# Excel files live under supporting_data/TCRMP_CVR/ (two levels up from this
# script: scripts/TCRMPcvr_makeAllPoints/ -> scripts/ -> repo_root).
# The orchestrator overrides both via positional argv[1] / argv[2].
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_SCRIPTS_DIR = os.path.dirname(SCRIPT_DIR)
_REPO_DIR = os.path.dirname(_SCRIPTS_DIR)
DEFAULT_CVR_ROOT = os.path.join(_REPO_DIR, 'supporting_data', 'TCRMP_CVR')
DEFAULT_OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TRANSECT_COLS_MODERN = {
    1: ('C', 'D', 'E'),
    2: ('F', 'G', 'H'),
    3: ('I', 'J', 'K'),
    4: ('L', 'M', 'N'),
    5: ('O', 'P', 'Q'),
    6: ('R', 'S', 'T'),
}

POINT_LABELS_20 = list('ABCDEFGHIJKLMNOPQRST')
POINT_LABELS_15 = list('ABCDEFGHIJKLMNO')
POINT_LABELS_10 = list('ABCDEFGHIJ')
ALL_POINT_LABELS = set(POINT_LABELS_20)

LOC_TO_SITE = {
    'flat cay': 'FLC', 'brewers bay': 'BWR', 'salt river west': 'SRW',
    'salt river east': 'SRE', 'cane bay': 'CBS', 'buck island stt': 'BIST',
    'buck island stx': 'BISX', 'long bay': 'LBH', 'south capella': 'SCP',
    'black point': 'BPT', 'botany bay': 'BTY', 'magens bay': 'MGN',
    'coculus rock': 'CRB',
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def excel_serial_to_date(serial):
    if isinstance(serial, datetime):
        return serial
    if isinstance(serial, (int, float)) and serial > 0:
        return datetime(1899, 12, 30) + timedelta(days=int(serial))
    return None


def parse_filename(filename):
    """Extract (date, site_code) from a CVR filename."""
    patterns = [
        (r'TCRMP(\d{8})_(?:CVR|CRV|CVR3D)_([A-Za-z]+)', 1, 2),
        (r'(\d{8})_TCRMP_CVR_([A-Za-z]+)', 1, 2),
        (r'TCRMP_(\d{8})_CVR_([A-Za-z]+)', 1, 2),
        (r'TCRMP_CVR_(\d{8})_([A-Za-z]+)', 1, 2),
    ]
    for pat, dg, sg in patterns:
        m = re.match(pat, filename)
        if m:
            try:
                return datetime.strptime(m.group(dg)[:8], '%Y%m%d'), m.group(sg).upper()
            except ValueError:
                pass
    # Fallback
    m = re.search(r'(\d{8,})', filename)
    if m:
        try:
            date = datetime.strptime(m.group(1)[:8], '%Y%m%d')
        except ValueError:
            date = None
        m2 = re.search(r'(?:CVR|CRV|CVR3D)_([A-Za-z]{2,5})', filename)
        site = m2.group(1).upper() if m2 else None
        return date, site
    return None, None


def detect_format(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        return 'legacy'
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        names_lower = [s.lower() for s in wb.sheetnames]
        wb.close()
        if 'rawdata' in names_lower:
            return 'modern'
        if 'raw_data' in names_lower:
            return 'intermediate'
        if any('tran' in n for n in names_lower):
            return 'legacy_xlsx'
        return 'unknown'
    except Exception:
        return 'unknown'


def find_sheet(wb, name_lower):
    for sn in wb.sheetnames:
        if sn.lower() == name_lower:
            return wb[sn]
    return None


# ---------------------------------------------------------------------------
# Codes parsers
# ---------------------------------------------------------------------------

def parse_codes_xlsx(wb):
    ws = find_sheet(wb, 'codes')
    if ws is None:
        return {}
    codes = {}
    first_data = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] and isinstance(row[0], str):
            if first_data is None:
                if row[0].upper() in ('CODE', 'CODES'):
                    continue
                first_data = True
            code = row[0].strip().upper()
            if len(row) >= 4 and row[3] is not None:
                category = str(row[2]).strip() if row[2] else ''
                meaning = str(row[3]).strip() if row[3] else ''
            else:
                category = str(row[1]).strip() if row[1] else ''
                meaning = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            codes[code] = {'category': category, 'name': meaning}
    return codes


def parse_codes_legacy(wb):
    codes = {}
    for sname in wb.sheet_names():
        if 'appendix' in sname.lower() or 'TRAN' in sname:
            ws = wb.sheet_by_name(sname)
            for r in range(ws.nrows):
                try:
                    val = ws.cell(r, 5).value
                    if isinstance(val, str) and '(' in val and ')' in val:
                        m = re.match(r'(.+?)\s*\((\w+)\)\s*-\s*(.+)', val.strip())
                        if m:
                            name = m.group(1).strip()
                            code = m.group(2).strip().upper()
                            category = m.group(3).strip().capitalize()
                            if code not in codes:
                                codes[code] = {'category': category, 'name': name}
                except (IndexError, TypeError):
                    continue
    return codes


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def parse_modern_rawdata(filepath, file_codes=None):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names_lower = {s.lower(): s for s in wb.sheetnames}

    if 'rawdata' in names_lower:
        raw_sheet = names_lower['rawdata']
    elif 'raw_data' in names_lower:
        raw_sheet = names_lower['raw_data']
    else:
        wb.close()
        return [], {}

    ws = wb[raw_sheet]
    rows_list = list(ws.rows)

    if len(rows_list) < 9:
        wb.close()
        return [], {}

    # Metadata
    location = None
    date = None

    try:
        c1 = rows_list[0][2].value
        if c1 and isinstance(c1, str):
            location = c1.strip()
    except (IndexError, TypeError):
        pass

    try:
        date = excel_serial_to_date(rows_list[2][2].value)
    except (IndexError, TypeError):
        pass

    # Codes
    codes = parse_codes_xlsx(wb)
    if not codes:
        codes = file_codes or {}

    # Point data (row 9+ = index 8+)
    points = []
    frame = 0

    for row_idx in range(8, len(rows_list)):
        row = rows_list[row_idx]

        try:
            a_val = row[0].value
            if a_val is not None and isinstance(a_val, (int, float)):
                frame = int(a_val)
        except (IndexError, TypeError):
            pass

        try:
            point_label = row[1].value
            if point_label is None:
                continue
            point_label = str(point_label).strip()
            if not point_label or point_label not in POINT_LABELS_20:
                continue
        except (IndexError, TypeError):
            continue

        for t_num, (pts_col, _, _) in TRANSECT_COLS_MODERN.items():
            col_idx = ord(pts_col) - ord('A')
            try:
                species = _clean_species(row[col_idx].value)
                if not species:
                    continue
                points.append({
                    'location': location,
                    'date': date,
                    'transect': t_num,
                    'frame': frame,
                    'point_label': point_label,
                    'species_code': species,
                })
            except (IndexError, TypeError):
                continue

    wb.close()
    return points, codes


def parse_legacy_xls(filepath):
    wb = xlrd.open_workbook(filepath)
    codes = parse_codes_legacy(wb)
    points = []

    for t_num in range(1, 7):
        sheet_name = f'{t_num:02d}TRAN'
        if sheet_name not in wb.sheet_names():
            continue
        ws = wb.sheet_by_name(sheet_name)

        location = None
        date = None
        try:
            location = ws.cell(0, 1).value
            if isinstance(location, str):
                location = location.strip()
        except (IndexError, TypeError):
            pass
        try:
            date = excel_serial_to_date(ws.cell(3, 1).value)
        except (IndexError, TypeError):
            pass

        # Skip the species lookup table at the top of each sheet.
        # The lookup table has species descriptions in column 5.
        # Find the last row with a description, then start at the next "A" label.
        last_desc_row = 9
        for r in range(10, ws.nrows):
            try:
                desc = ws.cell(r, 5).value
                if desc and str(desc).strip():
                    last_desc_row = r
            except (IndexError, TypeError):
                pass

        # Start at the next row after the lookup table that has label "A"
        data_start = last_desc_row + 1
        for r in range(last_desc_row + 1, ws.nrows):
            try:
                label = ws.cell(r, 1).value
                if label and str(label).strip().upper() == 'A':
                    data_start = r
                    break
            except (IndexError, TypeError):
                pass

        frame = 0
        for r in range(data_start, ws.nrows):
            try:
                point_label = ws.cell(r, 1).value
                if point_label is None or str(point_label).strip() == '':
                    continue
                point_label = str(point_label).strip().upper()
                if point_label not in ALL_POINT_LABELS:
                    continue

                species = None
                try:
                    species = _clean_species(ws.cell(r, 4).value)
                except (IndexError, TypeError):
                    pass
                if not species:
                    try:
                        species = _clean_species(ws.cell(r, 0).value)
                    except (IndexError, TypeError):
                        continue
                if not species:
                    continue

                # Frame increments when label resets to 'A'
                if point_label == 'A':
                    frame += 1
                points.append({
                    'location': location, 'date': date,
                    'transect': t_num, 'frame': frame,
                    'point_label': point_label, 'species_code': species,
                })
            except (IndexError, TypeError):
                continue

    return points, codes


def parse_legacy_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    codes = {}
    points = []

    for sname in wb.sheetnames:
        if 'appendix' in sname.lower():
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True):
                if row and len(row) > 5 and isinstance(row[5], str) and '(' in row[5]:
                    m = re.match(r'(.+?)\s*\((\w+)\)\s*-\s*(.+)', row[5].strip())
                    if m:
                        code = m.group(2).strip().upper()
                        codes[code] = {'category': m.group(3).strip().capitalize(), 'name': m.group(1).strip()}

    for t_num in range(1, 7):
        sheet_name = f'{t_num:02d}TRAN'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_list = list(ws.rows)

        location = None
        date = None
        try:
            location = rows_list[0][1].value
            if isinstance(location, str):
                location = location.strip()
        except (IndexError, TypeError):
            pass
        try:
            date = excel_serial_to_date(rows_list[3][1].value)
        except (IndexError, TypeError):
            pass

        # Skip species lookup table (rows with description in column 5)
        last_desc_row = 9
        for row_idx in range(10, len(rows_list)):
            row = rows_list[row_idx]
            try:
                desc = row[5].value if len(row) > 5 else None
                if desc and str(desc).strip():
                    last_desc_row = row_idx
            except (IndexError, TypeError):
                pass

        data_start = last_desc_row + 1
        for row_idx in range(last_desc_row + 1, len(rows_list)):
            row = rows_list[row_idx]
            try:
                label = row[1].value
                if label and str(label).strip().upper() == 'A':
                    data_start = row_idx
                    break
            except (IndexError, TypeError):
                pass

        frame = 0
        for row_idx in range(data_start, len(rows_list)):
            row = rows_list[row_idx]
            try:
                point_label = row[1].value
                if not point_label:
                    continue
                point_label = str(point_label).strip().upper()
                if point_label not in ALL_POINT_LABELS:
                    continue

                species = None
                try:
                    species = _clean_species(row[4].value)
                except (IndexError, TypeError):
                    pass
                if not species:
                    try:
                        species = _clean_species(row[0].value)
                    except (IndexError, TypeError):
                        continue
                if not species:
                    continue

                # Frame increments when label resets to 'A'
                if point_label == 'A':
                    frame += 1
                points.append({
                    'location': location, 'date': date,
                    'transect': t_num, 'frame': frame,
                    'point_label': point_label, 'species_code': species,
                })
            except (IndexError, TypeError):
                continue

    wb.close()
    return points, codes


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def find_all_cvr_files(root):
    files = []
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn.startswith('~') or fn.startswith('.'):
                continue
            ext = os.path.splitext(fn)[1].lower()
            if ext in ('.xlsx', '.xls'):
                files.append(os.path.join(dirpath, fn))
    return sorted(files)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(cvr_root=None, output_dir=None):
    if cvr_root is None:
        cvr_root = os.path.abspath(DEFAULT_CVR_ROOT)
    if output_dir is None:
        output_dir = os.path.abspath(DEFAULT_OUTPUT_DIR)

    os.makedirs(output_dir, exist_ok=True)

    log_lines = []

    def log(msg):
        print(msg)
        log_lines.append(msg)

    files = find_all_cvr_files(cvr_root)
    log(f"Found {len(files)} Excel files to parse in {cvr_root}")

    all_points = []
    master_codes = {}
    errors = []
    stats = defaultdict(int)

    for i, filepath in enumerate(files):
        filename = os.path.basename(filepath)
        if (i + 1) % 50 == 0 or i == 0:
            log(f"  Processing file {i+1}/{len(files)}: {filename}")

        file_date, file_site = parse_filename(filename)

        try:
            fmt = detect_format(filepath)
            stats[fmt] += 1

            if fmt in ('modern', 'intermediate'):
                points, codes = parse_modern_rawdata(filepath)
            elif fmt == 'legacy':
                points, codes = parse_legacy_xls(filepath)
            elif fmt == 'legacy_xlsx':
                points, codes = parse_legacy_xlsx(filepath)
            else:
                errors.append((filepath, f"Unknown format: {fmt}"))
                continue

            master_codes.update(codes)

            for pt in points:
                if pt['date'] is None:
                    pt['date'] = file_date
                if file_site:
                    pt['site'] = file_site
                else:
                    pt['site'] = LOC_TO_SITE.get(
                        pt.get('location', '').lower().strip() if pt.get('location') else '')

            all_points.extend(points)

        except Exception as e:
            errors.append((filepath, str(e)))
            continue

    log(f"\nParsed {len(all_points)} total points from {len(files) - len(errors)} files")
    log(f"Format distribution: {dict(stats)}")
    if errors:
        log(f"\n{len(errors)} files had errors:")
        for fp, err in errors[:20]:
            log(f"  {os.path.basename(fp)}: {err}")
        if len(errors) > 20:
            log(f"  ... and {len(errors) - 20} more")

    # Build dataframe
    df = pd.DataFrame(all_points)
    if len(df) == 0:
        log("ERROR: No data parsed!")
        return

    df['species_name'] = df['species_code'].map(
        lambda c: master_codes.get(c, {}).get('name', ''))
    df['category'] = df['species_code'].map(
        lambda c: master_codes.get(c, {}).get('category', ''))
    df['year'] = df['date'].apply(lambda d: d.year if d else None)

    df = df[['date', 'year', 'site', 'transect', 'frame', 'point_label',
             'species_code', 'species_name', 'category']]

    # Save with date stamp
    datestamp = datetime.now().strftime('%Y%m%d')
    out_path = os.path.join(output_dir, f'all_points_{datestamp}.csv')
    df.to_csv(out_path, index=False)
    log(f"\nSaved {len(df)} points to {out_path}")

    # Master codes
    codes_df = pd.DataFrame([
        {'code': k, 'category': v['category'], 'name': v['name']}
        for k, v in sorted(master_codes.items())
    ])
    codes_path = os.path.join(output_dir, 'master_codes.csv')
    codes_df.to_csv(codes_path, index=False)
    log(f"Saved {len(codes_df)} species codes to {codes_path}")

    # Summary
    log(f"\nSummary:")
    log(f"  Years: {sorted(df['year'].dropna().unique().astype(int))}")
    log(f"  Sites: {sorted(df['site'].dropna().unique())}")
    log(f"  Unique species codes: {df['species_code'].nunique()}")
    log(f"  Points with known category: {(df['category'] != '').sum()}/{len(df)}")

    # Save log
    log_path = os.path.join(output_dir, 'parse_log.txt')
    with open(log_path, 'w') as f:
        f.write('\n'.join(log_lines))
    print(f"Saved log to {log_path}")


if __name__ == '__main__':
    cvr = sys.argv[1] if len(sys.argv) > 1 else None
    out = sys.argv[2] if len(sys.argv) > 2 else None
    main(cvr, out)
